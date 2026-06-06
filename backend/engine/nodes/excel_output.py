"""Build a multi-tab .xlsx and return base64 bytes."""
from __future__ import annotations

from pathlib import Path
from typing import Any

from ..context import RunContext
from ..node_spec import _spec_from_yaml

_HERE = Path(__file__).parent
  
import base64
import io

from ..output_files import write_export_file


def _normalize_tab_names(raw: Any) -> list[str]:
    """Accept tabNames as comma-separated string (Studio YAML) or list (Copilot JSON)."""
    if raw is None:
        return []
    if isinstance(raw, list):
        return [str(n).strip() for n in raw if str(n).strip()]
    if isinstance(raw, str):
        return [n.strip() for n in raw.split(",") if n.strip()]
    text = str(raw).strip()
    return [text] if text else []


def run(node: dict, ctx: RunContext, incoming: dict[str, Any]) -> dict[str, Any]:
    cfg = node.get("config") or {}
    tab_names = _normalize_tab_names(cfg.get("tabNames"))
    filename = cfg.get("filename") or "output.xlsx"

    datasets: list[dict] = []
    idx = 0
    for up in incoming.values():
        if isinstance(up, dict) and isinstance(up.get("rows"), list):
            name = tab_names[idx] if idx < len(tab_names) else f"Sheet{idx + 1}"
            datasets.append({"tab": name[:31], "rows": up["rows"]})
            idx += 1
    if not datasets:
        return {"filename": filename, "tabs": 0, "rowsWritten": 0, "note": "No rows received"}

    highlight_col = cfg.get("highlightColumn")
    highlight_map_str = cfg.get("highlightMapping") or ""
    
    color_map = {
        "red": "FFCDD2",      # Soft red
        "yellow": "FFF9C4",   # Soft yellow
        "blue": "BBDEFB",     # Soft blue
        "green": "C8E6C9",    # Soft green
    }
    
    highlights = {}
    if highlight_col and highlight_map_str:
        for pair in highlight_map_str.split(","):
            if ":" in pair:
                val, color = pair.split(":", 1)
                val = val.strip().lower()
                color = color.strip().lower()
                hex_color = color_map.get(color, color.replace("#", "").upper())
                highlights[val] = hex_color

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
    except Exception as exc:
        return {"filename": filename, "error": f"openpyxl missing: {exc}", "tabs": 0, "rowsWritten": 0}

    wb = Workbook()
    wb.remove(wb.active)  # default sheet
    total = 0
    preview = []
    for ds in datasets:
        ws = wb.create_sheet(title=ds["tab"])
        rows = ds["rows"]
        if rows:
            cols = list(rows[0].keys())
            ws.append(cols)
            
            col_idx = -1
            if highlight_col in cols:
                col_idx = cols.index(highlight_col)
                
            for r in rows:
                row_vals = [r.get(c) for c in cols]
                ws.append(row_vals)
                
                if col_idx != -1:
                    cell_val = str(row_vals[col_idx]).strip().lower()
                    if cell_val in highlights:
                        hex_color = highlights[cell_val]
                        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                        for cell in ws[ws.max_row]:
                            cell.fill = fill
        total += len(rows)
        preview.append({"tab": ds["tab"], "rows": rows[:3], "total": len(rows)})

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    path, download_url = write_export_file(filename, data, default_name="output.xlsx")
    ctx.report_path = str(path)
    return {
        "filename": path.name,
        "tabs": len(datasets),
        "tabNames": [d["tab"] for d in datasets],
        "rowsWritten": total,
        "byteSize": len(data),
        "base64": base64.b64encode(data).decode("ascii"),
        "preview": preview,
        "report_path": str(path),
        "download_url": download_url,
    }
  
NODE_SPEC = _spec_from_yaml(_HERE / "excel_output.yaml", run)
  