"""
Phase 7 — FISL (Fixed Income Spoof / Layering) golden-path test.

This is the chassis-portability proof: a brand-new scenario built
*entirely* on Phase 1-5 primitives with zero new node code. The
workflow exercises the nested-MAP pattern (outer over venue
"spoof instances", inner over book) — falsifiable evidence that
control-flow primitives compose recursively.
"""
from __future__ import annotations

import json
import os

import pytest
from openpyxl import load_workbook

from engine.dag_runner import run_workflow
from engine.nodes import consolidated_summary as cs
from engine.nodes import section_summary as ss
from engine.validator import validate_dag


WORKFLOW_PATH = os.path.join(
    os.path.dirname(__file__), "..", "workflows", "fisl_workflow.json"
)


@pytest.fixture
def mocked_llm(monkeypatch):
    def fake_section(prompt: str) -> str:
        if "{" in prompt and "}" in prompt:
            try:
                start = prompt.index("{")
                depth = 0
                end = start
                for i in range(start, len(prompt)):
                    if prompt[i] == "{":
                        depth += 1
                    elif prompt[i] == "}":
                        depth -= 1
                        if depth == 0:
                            end = i + 1
                            break
                blob = prompt[start:end]
                facts = json.loads(blob)
                return "Findings: " + ", ".join(
                    f"{k}={v}" for k, v in facts.items()
                )
            except Exception:
                pass
        return "Findings narrative."

    monkeypatch.setattr(ss, "_llm_narrative", fake_section)
    monkeypatch.setattr(cs, "_llm_summary", lambda p: "Executive summary text.")


def test_fisl_workflow_validates_clean():
    with open(WORKFLOW_PATH) as f:
        dag = json.load(f)
    result = validate_dag(dag)
    assert result.valid, [i.message for i in result.errors]
    assert result.warnings == []


def test_fisl_workflow_runs_end_to_end_with_nested_map(tmp_path, mocked_llm, monkeypatch):
    monkeypatch.setenv("DBSHERPA_OUTPUT_DIR", str(tmp_path))

    with open(WORKFLOW_PATH) as f:
        dag = json.load(f)

    alert = {
        "trader_id": "T042",
        "alert_id": "FISL-2024-0007",
        "alert_date": "2024-01-15",
        "event_time": "2024-01-15 09:00",
    }

    ctx = run_workflow(dag, alert)

    # Window built — keeps the chassis windowing chain intact.
    window = ctx.get("window")
    assert window["start_time"].startswith("2024-01-15T08:45")
    assert window["end_time"].startswith("2024-01-15T09:15")

    # Outer GROUP_BY: per-venue datasets and venue keys
    venue_keys = ctx.get("venue_keys")["values"]
    assert len(venue_keys) >= 2  # mock orders span EBS/Reuters/Bloomberg
    for v in venue_keys:
        assert f"orders_by_venue_{v}" in ctx.datasets

    # Outer MAP harvested per-venue results
    per_instr = ctx.get("per_instrument")
    assert per_instr and set(per_instr["results"].keys()) == set(venue_keys)

    # Inner MAP ran inside each outer iteration; per_book is harvested back
    for v in venue_keys:
        venue_result = per_instr["results"][v]
        assert "per_book" in venue_result, f"inner MAP results missing for venue {v}"
        assert "instance_qty_total" in venue_result
        per_book = venue_result["per_book"]
        assert "results" in per_book
        # At least one book per venue (mock data has two books)
        assert len(per_book["results"]) >= 1
        for book_key, book_vals in per_book["results"].items():
            assert "book_qty" in book_vals
            assert "book_order_count" in book_vals

        # Outer MAP also harvested the iteration's `orders` dataset
        assert f"per_instrument_{v}_orders" in ctx.datasets

    # FEATURE_ENGINE produced ladder + intermediate ladder_long
    assert "ladder" in ctx.datasets
    assert "ladder_long" in ctx.datasets
    ladder = ctx.datasets["ladder"]
    assert "minute_bucket" in ladder.columns

    # SIGNAL_CALCULATOR (SPOOFING) — 5-column contract
    sigs = ctx.datasets["orders_signals"]
    for col in ("_signal_flag", "_signal_score", "_signal_reason", "_signal_type", "_signal_window"):
        assert col in sigs.columns
    assert (sigs["_signal_type"] == "SPOOFING").all()

    # DATA_HIGHLIGHTER coloured the orders
    assert "_highlight_colour" in ctx.datasets["orders_signals_highlighted"].columns

    # DECISION_RULE — rules-mode disposition + severity
    assert ctx.disposition in ("ESCALATE", "REVIEW", "DISMISS")
    assert ctx.get("severity") in ("CRITICAL", "HIGH", "MEDIUM", "LOW")

    # All three SECTION_SUMMARY modes
    assert "orders_overview" in ctx.sections        # fact_pack_llm
    assert "spoof_signals" in ctx.sections          # templated

    # Executive summary populated
    assert ctx.executive_summary == "Executive summary text."

    # Report file written with static + per-venue tabs
    assert os.path.isfile(ctx.report_path)
    wb = load_workbook(ctx.report_path)
    sheets = set(wb.sheetnames)
    assert "All Orders" in sheets
    assert "Order-Book Ladder" in sheets
    assert "Spoof Signals" in sheets
    for v in venue_keys:
        expected = f"Venue {v}"[:31]
        assert expected in sheets, f"missing per-venue tab '{expected}' in {sheets}"
