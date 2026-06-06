"""
Phase 5 — comms keyword categories + hits filter, REPORT_OUTPUT
REPORT_OUTPUT.expand_from, recursive MAP sub_workflow validation.
"""
from __future__ import annotations

import os
import tempfile

import pandas as pd

from engine.context import RunContext
from engine.nodes.comms_collector import handle_comms_collector
from engine.nodes.report_output import handle_report_output
from engine.validator import validate_dag


# ---------------------------------------------------------------------------
# Comms: keyword_categories + emit_hits_only
# ---------------------------------------------------------------------------
def test_comms_keyword_categories_adds_per_cat_columns():
    ctx = RunContext()
    handle_comms_collector({"config": {
        "query_template": "*:*",
        "keywords": [],
        "keyword_categories": {
            "INTENT": ["layering", "spoofing", "front-run"],
            "TIMING": ["fix", "WM fix", "benchmark"],
        },
        "emit_hits_only": True,
        "output_name": "comms",
    }}, ctx)

    df = ctx.datasets["comms"]
    assert "_matched_categories" in df.columns
    assert "_hit_INTENT" in df.columns
    assert "_hit_TIMING" in df.columns
    # Every row with INTENT in _matched_categories must have _hit_INTENT True
    for _, row in df.iterrows():
        assert row["_hit_INTENT"] == ("INTENT" in row["_matched_categories"])

    # emit_hits_only: filtered dataset published
    assert "comms_hits" in ctx.datasets
    hits = ctx.datasets["comms_hits"]
    assert len(hits) == int(df["_keyword_hit"].sum())
    assert ctx.get("comms_hits_count") == len(hits)


def test_comms_keyword_categories_optional_noop():
    ctx = RunContext()
    handle_comms_collector({"config": {
        "query_template": "*:*",
        "output_name": "comms",
    }}, ctx)
    df = ctx.datasets["comms"]
    assert "_matched_categories" not in df.columns
    assert "comms_hits" not in ctx.datasets


# ---------------------------------------------------------------------------
# REPORT_OUTPUT: expand_from
# ---------------------------------------------------------------------------
def test_report_output_expands_tabs_from_map_result(tmp_path):
    """Iterating MAP's `{results: {...}}` dict — keys become tabs."""
    ctx = RunContext()
    ctx.datasets["per_book_A_orders"] = pd.DataFrame({"x": [1, 2]})
    ctx.datasets["per_book_B_orders"] = pd.DataFrame({"x": [3]})
    ctx.set("per_book", {"results": {"A": {}, "B": {}}})

    out = str(tmp_path / "r.xlsx")
    handle_report_output({"config": {
        "output_path": out,
        "tabs": [{
            "expand_from": "{context.per_book}",
            "as": "key",
            "name": "Book {key}",
            "dataset": "per_book_{key}_orders",
            "include_highlights": False,
        }],
    }}, ctx)

    assert os.path.isfile(out)
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert "Book A" in wb.sheetnames
    assert "Book B" in wb.sheetnames


def test_report_output_expand_from_skips_missing(tmp_path):
    """Items whose target dataset doesn't exist are silently skipped."""
    ctx = RunContext()
    ctx.datasets["per_book_A_orders"] = pd.DataFrame({"x": [1]})
    ctx.set("per_book", {"results": {"A": {}, "B": {}}})
    out = str(tmp_path / "r.xlsx")
    handle_report_output({"config": {
        "output_path": out,
        "tabs": [{
            "expand_from": "{context.per_book}",
            "as": "key",
            "name": "Book {key}",
            "dataset": "per_book_{key}_orders",
        }],
    }}, ctx)
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert "Book A" in wb.sheetnames
    assert "Book B" not in wb.sheetnames


def test_report_output_expand_from_context_list(tmp_path):
    """The new general path: expand from a plain list in ctx.values."""
    ctx = RunContext()
    ctx.datasets["execs_FX-SPOT"] = pd.DataFrame({"qty": [10]})
    ctx.datasets["execs_FX-FWD"]  = pd.DataFrame({"qty": [20]})
    ctx.set("books", ["FX-SPOT", "FX-FWD"])

    out = str(tmp_path / "r.xlsx")
    handle_report_output({"config": {
        "output_path": out,
        "tabs": [{
            "expand_from": "{context.books}",
            "as": "book",
            "name": "Executions · {book}",
            "dataset": "execs_{book}",
        }],
    }}, ctx)

    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert "Executions · FX-SPOT" in wb.sheetnames
    assert "Executions · FX-FWD" in wb.sheetnames


# ---------------------------------------------------------------------------
# Validator: recursive MAP sub_workflow check
# ---------------------------------------------------------------------------
def _base_dag_with_map(sub_workflow: dict) -> dict:
    return {
        "workflow_id": "w",
        "nodes": [
            {"id": "n01", "type": "ALERT_TRIGGER", "label": "start", "config": {}},
            {"id": "n02", "type": "EXECUTION_DATA_COLLECTOR", "label": "orders", "config": {
                "source": "hs_client_order", "query_template": "*:*", "output_name": "orders",
            }},
            {"id": "n03", "type": "GROUP_BY", "label": "group", "config": {
                "input_name": "orders", "group_by_column": "book",
                "output_prefix": "obb",
            }},
            {"id": "n04", "type": "MAP", "label": "per_book", "config": {
                "keys_key": "orders_keys",
                "iteration_ctx_key": "current_book",
                "sub_workflow": sub_workflow,
                "output_name": "per_book",
            }},
            {"id": "n05", "type": "REPORT_OUTPUT", "label": "report", "config": {
                "output_path": "output/r.xlsx",
            }},
        ],
        "edges": [
            {"from": "n01", "to": "n02"},
            {"from": "n02", "to": "n03"},
            {"from": "n03", "to": "n04"},
            {"from": "n04", "to": "n05"},
        ],
    }


def test_validator_flags_unknown_node_in_map_sub_workflow():
    sub = {
        "nodes": [{"id": "s1", "type": "NO_SUCH_NODE", "config": {}, "label": "?"}],
        "edges": [],
    }
    result = validate_dag(_base_dag_with_map(sub))
    errs = [i for i in result.errors if i.node_id == "n04"]
    assert any("sub_workflow" in i.message.lower() and "NO_SUCH_NODE" in i.message for i in errs)


def test_validator_flags_missing_required_param_in_sub_workflow():
    sub = {
        "nodes": [{
            "id": "s1", "type": "EXTRACT_SCALAR", "label": "x",
            "config": {"input_name": "orders", "reducer": "sum"},  # column missing
        }],
        "edges": [],
    }
    result = validate_dag(_base_dag_with_map(sub))
    errs = [i for i in result.errors if i.node_id == "n04"]
    assert any("column" in i.message.lower() and "sub_workflow" in i.message.lower() for i in errs)


def test_validator_accepts_valid_map_sub_workflow():
    sub = {
        "nodes": [{
            "id": "s1", "type": "EXTRACT_SCALAR", "label": "qty",
            "config": {"input_name": "orders", "column": "quantity",
                       "reducer": "sum", "output_name": "book_qty"},
        }],
        "edges": [],
    }
    result = validate_dag(_base_dag_with_map(sub))
    # No sub-workflow errors — MAP itself should be clean.
    sub_errs = [i for i in result.errors if i.node_id == "n04" and "sub_workflow" in (i.message or "").lower()]
    assert sub_errs == []


def test_validator_flags_empty_sub_workflow():
    sub = {"nodes": [], "edges": []}
    result = validate_dag(_base_dag_with_map(sub))
    assert any(i.node_id == "n04" and "sub_workflow" in (i.message or "").lower() for i in result.errors)
