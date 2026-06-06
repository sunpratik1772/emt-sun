"""
Phase 6 — FRO chassis-v2 reference workflow golden-path test.

Validates and runs `workflows/fx_fro_v2_workflow.json` end-to-end with
the LLM seam stubbed. Confirms:

  * The workflow validates clean (no errors, no warnings).
  * Topological execution succeeds with no node contract violations.
  * All chassis primitives wire correctly: TIME_WINDOW publishes a
    window, every collector honours it, GROUP_BY partitions orders,
    MAP fans out per book and harvests per-book values + datasets,
    SECTION_SUMMARY runs in fact_pack_llm and event_narrative modes,
    REPORT_OUTPUT writes one tab per book via tabs_from_map.
  * The Excel report file is written and contains the expected tabs.
"""
from __future__ import annotations

import json
import os

import pytest
from openpyxl import load_workbook

from engine.dag_runner import run_workflow
from engine.nodes import consolidated_summary as cs
from engine.nodes import section_summary as ss
from engine.validator import validate_dag


WORKFLOW_PATH = os.path.join(
    os.path.dirname(__file__), "..", "workflows", "fx_fro_v2_workflow.json"
)


@pytest.fixture
def mocked_llm(monkeypatch):
    """Stub the section + consolidated LLM seams with deterministic prose
    that includes the required fact values so fact_pack_llm doesn't retry."""
    def fake_section(prompt: str) -> str:
        # Pull the JSON facts out of the prompt and stitch every value
        # back into the response so required-fact validation passes.
        if "{" in prompt and "}" in prompt:
            try:
                start = prompt.index("{")
                # naive bracket scan to find the matching brace
                depth = 0
                end = start
                for i in range(start, len(prompt)):
                    if prompt[i] == "{":
                        depth += 1
                    elif prompt[i] == "}":
                        depth -= 1
                        if depth == 0:
                            end = i + 1
                            break
                blob = prompt[start:end]
                facts = json.loads(blob)
                return "Findings: " + ", ".join(
                    f"{k}={v}" for k, v in facts.items()
                )
            except Exception:
                pass
        return "Findings narrative."

    def fake_exec(prompt: str) -> str:
        return "Executive summary text."

    monkeypatch.setattr(ss, "_llm_narrative", fake_section)
    monkeypatch.setattr(cs, "_llm_summary", fake_exec)


def test_fro_v2_workflow_validates_clean():
    with open(WORKFLOW_PATH) as f:
        dag = json.load(f)
    result = validate_dag(dag)
    assert result.valid, [i.message for i in result.errors]
    assert result.warnings == []


def test_fro_v2_workflow_runs_end_to_end(tmp_path, mocked_llm, monkeypatch):
    monkeypatch.setenv("DBSHERPA_OUTPUT_DIR", str(tmp_path))

    with open(WORKFLOW_PATH) as f:
        dag = json.load(f)

    alert = {
        "trader_id": "T001",
        "currency_pair": "EUR/USD",
        "alert_id": "FRO-2024-0042",
        "alert_date": "2024-01-15",
        "event_time": "2024-01-15 09:00",
    }
    # Seed alert fields into ctx via the standard ALERT_TRIGGER path.
    # The ALERT_TRIGGER handler reads from alert_payload, so populating
    # alert_payload is sufficient.
    ctx = run_workflow(dag, alert)

    # Window built by TIME_WINDOW
    window = ctx.get("window")
    assert window["start_time"].startswith("2024-01-15T08:30")
    assert window["end_time"].startswith("2024-01-15T09:30")
    assert window["buffer_minutes"] == {"pre": 30, "post": 30}

    # Collectors emitted their datasets
    assert "orders" in ctx.datasets
    assert "executions" in ctx.datasets
    assert "comms" in ctx.datasets
    assert "comms_hits" in ctx.datasets        # emit_hits_only
    assert "market" in ctx.datasets

    # Comms keyword categories applied
    comms = ctx.datasets["comms"]
    for col in ("_matched_categories", "_hit_INTENT", "_hit_TIMING", "_hit_COVERUP"):
        assert col in comms.columns

    # GROUP_BY partitioned orders by book
    assert ctx.get("book_keys") and ctx.get("book_keys")["values"]
    book_keys = ctx.get("book_keys")["values"]
    for k in book_keys:
        assert f"orders_by_book_{k}" in ctx.datasets

    # MAP fanned out per book
    per_book = ctx.get("per_book")
    assert per_book and set(per_book["results"].keys()) == set(book_keys)
    for k in book_keys:
        result_for_k = per_book["results"][k]
        assert "book_qty_total" in result_for_k
        assert "book_order_count" in result_for_k
        # Aliased dataset harvested back to parent under per_book_<k>_orders
        assert f"per_book_{k}_orders" in ctx.datasets

    # FEATURE_ENGINE published bucketed + derived columns
    feats = ctx.datasets["executions_features"]
    assert "minute_bucket" in feats.columns
    assert "notional_per_unit" in feats.columns

    # SIGNAL_CALCULATOR (FRONT_RUNNING) emitted the 5-column contract
    sigs = ctx.datasets["executions_signals"]
    for col in ("_signal_flag", "_signal_score", "_signal_reason", "_signal_type", "_signal_window"):
        assert col in sigs.columns
    assert (sigs["_signal_type"] == "FRONT_RUNNING").all()

    # DATA_HIGHLIGHTER coloured the rows
    hl = ctx.datasets["executions_signals_highlighted"]
    assert "_highlight_colour" in hl.columns

    # DECISION_RULE produced disposition + severity + score + matched_rule
    assert ctx.disposition in ("ESCALATE", "REVIEW", "DISMISS")
    assert ctx.get("severity") in ("CRITICAL", "HIGH", "MEDIUM", "LOW")
    assert isinstance(ctx.get("score"), float)
    assert isinstance(ctx.get("matched_rule"), str)

    # All three SECTION_SUMMARY modes rendered
    assert "executions_analysis" in ctx.sections
    assert "comms_narrative" in ctx.sections
    assert "signals_overview" in ctx.sections
    exec_section = ctx.sections["executions_analysis"]
    # Required facts present in the (mocked) narrative
    assert "exec_count" in exec_section["narrative"]
    assert exec_section["stats"]["row_count"] > 0

    # Executive summary populated
    assert ctx.executive_summary == "Executive summary text."

    # Report file written and contains static + dynamic tabs
    assert os.path.isfile(ctx.report_path)
    wb = load_workbook(ctx.report_path)
    sheets = set(wb.sheetnames)
    assert "All Orders" in sheets
    assert "Executions" in sheets
    assert "Signal Detections" in sheets
    assert "Comms Hits" in sheets
    assert "Market" in sheets
    for k in book_keys:
        # name_template = "Book {key} Orders" → truncated to 31 chars
        expected = f"Book {k} Orders"[:31]
        assert expected in sheets
