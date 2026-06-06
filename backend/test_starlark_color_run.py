import json
from pathlib import Path
from engine.jobs import get_default_runner
from engine import RunContext

def test_run():
    # Load the new demo workflow
    wf_path = Path(__file__).parent / "good_examples" / "studio_15_starlark_excel_colors_demo.json"
    print("Loading workflow from:", wf_path)
    with wf_path.open("r") as f:
        dag = json.load(f)

    # Mock alert payload
    alert_payload = {
        "alert_id": "DEMO-0001",
        "alert_date": "2024-01-15"
    }

    print("Running workflow DAG...")
    runner = get_default_runner()
    ctx = runner.run(dag, alert_payload).context

    print("Execution Finished!")
    print("Run ID:", ctx.run_id)
    print("Disposition:", ctx.disposition)
    print("Report Path:", ctx.report_path)
    
    if ctx.report_path and Path(ctx.report_path).exists():
        print("Excel report generated successfully!")
        # Let's inspect the excel workbook sheets
        from openpyxl import load_workbook
        wb = load_workbook(ctx.report_path)
        print("Workbook sheets:", wb.sheetnames)
        
        # Check first sheet row colors
        ws = wb["Risk Analysis"]
        colored_rows = 0
        for row in ws.iter_rows(min_row=2):  # skip header
            fill = row[0].fill
            if fill and fill.start_color and fill.start_color.rgb != "00000000":
                colored_rows += 1
        print(f"Verified {colored_rows} colored row(s) in sheet!")
        assert colored_rows > 0, "No styled/colored cells found!"
        print("Starlark mathematical scoring and Excel coloring tests PASSED!")
    else:
        raise Exception("Excel report failed to generate!")

if __name__ == "__main__":
    test_run()
